# Importing different modules
from msilib.schema import Font
from platform import python_version
# A Python library to read/write Excel file
import openpyxl as op
# For changing the style of a particular cell like changing colours, fonts or apply borders
from openpyxl.styles import PatternFill, Font

try:
    # Loading the given input file in a Workbook
    wb = op.load_workbook('input_octant_transition_identify.xlsx')
    # Above line can have error if input file has wrong format
    # other than .xlsx

except Exception as e:  # Exception = openpyxl.utils.exceptions.InvalidFileException
    print("There was some error due to " + str(e))


try:
    # Selecting the current active sheet
    sheet = wb.active
    # Above line can have error if input file is read only workbook
    # and we are trying to modify it

except Exception as e:  # Exception = openpyxl.utils.exceptions.ReadOnlyWorkbookException
    print("There was some error due to " + str(e))


sheet['E1'] = "U_Avg"  
sheet['F1'] = "V_Avg"
sheet['G1'] = "W_Avg"

# Total rows
row_count = sheet.max_row

# Calculating average by the formula
sheet['E2'] = '= AVERAGE(B2:B' + str(row_count) + ')'
sheet['F2'] = '= AVERAGE(C2:C' + str(row_count) + ')'
sheet['G2'] = '= AVERAGE(D2:D' + str(row_count) + ')'

sheet['H1'] = "U' = U - U_Avg"  
sheet['I1'] = "V' = V - V_Avg"
sheet['J1'] = "W' = W - W_Avg"

# Calculated respective velocities for entire data
for i in range(2, row_count + 1):
    sheet['H' + str(i)] = '= B' + str(i) + ' - E2'
    sheet['I' + str(i)] = '= C{} - F2'.format(i)
    sheet['J' + str(i)] = '= D{} - G2'.format(i)

# Creating Octant Column
sheet['K1'] = "Octant"

for i in range(2, row_count + 1):
    # Direct formula for Counting a particular occurence according to the conditions
    sheet[f'K{i}'] = f'''= IF( AND(H{i}>=0, I{i}>=0, J{i}>=0), "+1", IF( AND(H{i}>=0, I{i}>=0, J{i}<0), "-1",
    IF( AND(H{i}<0, I{i}>=0, J{i}>0), "+2", IF( AND(H{i}<0, I{i}>=0, J{i}<0), "-2", IF( AND(H{i}<0, I{i}<0, J{i}>0), "+3",
    IF( AND(H{i}<0, I{i}<0, J{i}<0), "-3", IF( AND(H{i}>=0, I{i}<0, J{i}>0), "+4", IF( AND(H{i}>=0, I{i}<0, J{i}<0), "-4"))))))))'''

def octant_transition_count(mod = 5000):
    # Assigning respective cell value
    sheet['L3'] = 'User Input'
    # Fills the selected colour in the chosen cell
    sheet['L3'].fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

    sheet['M2'] = 'Overall Count'

    sheet['M3'] = 'Mod ' + str(mod)
    # Fills the selected colour in the chosen cell
    sheet['M3'].fill = PatternFill(start_color='BBECCD', end_color='BBECCD', fill_type='solid')
    # Changes the font colour with the desired one
    sheet['M3'].font = Font(color='005C29')
    
    sheet['M13'] = 'Overall Transition Count'
    sheet['M15'] = 'Count'
    sheet['L16'] = 'From'
    sheet['N14'] = 'To'

    sheet['N1'] = sheet['N15'] = sheet['M16'] = '+1'
    sheet['O1'] = sheet['O15'] = sheet['M17'] = '-1'
    sheet['P1'] = sheet['P15'] = sheet['M18'] = '+2'
    sheet['Q1'] = sheet['Q15'] = sheet['M19'] = '-2'
    sheet['R1'] = sheet['R15'] = sheet['M20'] = '+3'
    sheet['S1'] = sheet['S15'] = sheet['M21'] = '-3'
    sheet['T1'] = sheet['T15'] = sheet['M22'] = '+4'
    sheet['U1'] = sheet['U15'] = sheet['M23'] = '-4'

    # Direct formula to find the total count of a particular octant value
    sheet['N2'] = f'= COUNTIF($K$2:$K${row_count},N1)'
    sheet['O2'] = f'= COUNTIF($K$2:$K${row_count},O1)'
    sheet['P2'] = f'= COUNTIF($K$2:$K${row_count},P1)'
    sheet['Q2'] = f'= COUNTIF($K$2:$K${row_count},Q1)'
    sheet['R2'] = f'= COUNTIF($K$2:$K${row_count},R1)'
    sheet['S2'] = f'= COUNTIF($K$2:$K${row_count},S1)'
    sheet['T2'] = f'= COUNTIF($K$2:$K${row_count},T1)'
    sheet['U2'] = f'= COUNTIF($K$2:$K${row_count},U1)'


    try:
        # Creating a temporary column containg one next values of octant
        sheet['AM1'] = "Temp_Column"
        for i in range(row_count):
            sheet[f'AM{i + 2}'] = f'= K{i + 3}'

    except:  # Exception = Unknown
        print("An Unknown error ocurred")
    

    # Creating List of cols value which will be used later
    cols = ['+1', '-1', '+2', '-2', '+3', '-3', '+4', '-4']

    # No of iterations for the below loop
    iter = row_count//mod

    for i in range(iter + 1):
        # start = starting value of a given range
        # end = ending value of a given range
        start = i*mod
        if(i == iter):
            end = row_count
        else:
            end = (i + 1)*mod - 1
        sheet[f'M{4 + i}'] = f'{start} - {end}'

        # Direct formula to find the total count of a particular octant value in given range
        # from start + 2 to end + 2
        sheet[f'N{4 + i}'] = f'= COUNTIF($K${start + 2}:$K${end + 2},N1)'
        sheet[f'O{4 + i}'] = f'= COUNTIF($K${start + 2}:$K${end + 2},O1)'
        sheet[f'P{4 + i}'] = f'= COUNTIF($K${start + 2}:$K${end + 2},P1)'
        sheet[f'Q{4 + i}'] = f'= COUNTIF($K${start + 2}:$K${end + 2},Q1)'
        sheet[f'R{4 + i}'] = f'= COUNTIF($K${start + 2}:$K${end + 2},R1)'
        sheet[f'S{4 + i}'] = f'= COUNTIF($K${start + 2}:$K${end + 2},S1)'
        sheet[f'T{4 + i}'] = f'= COUNTIF($K${start + 2}:$K${end + 2},T1)'
        sheet[f'U{4 + i}'] = f'= COUNTIF($K${start + 2}:$K${end + 2},U1)'

        sheet[f'M{5 + iter}'] = 'Verified'
        # Adding verified row by finding sumof the respective columns
        sheet[f'N{5 + iter}'] = f'= SUM(N4:N{4 + iter})'
        sheet[f'O{5 + iter}'] = f'= SUM(O4:O{4 + iter})'
        sheet[f'P{5 + iter}'] = f'= SUM(P4:P{4 + iter})'
        sheet[f'Q{5 + iter}'] = f'= SUM(Q4:Q{4 + iter})'
        sheet[f'R{5 + iter}'] = f'= SUM(R4:R{4 + iter})'
        sheet[f'S{5 + iter}'] = f'= SUM(S4:S{4 + iter})'
        sheet[f'T{5 + iter}'] = f'= SUM(T4:T{4 + iter})'
        sheet[f'U{5 + iter}'] = f'= SUM(U4:U{4 + iter})'

        # Assigning cell values for Transition count tables for each iteration
        sheet[f'L{30 + i*13}'] = 'From'
        sheet[f'M{27 + i*13}'] = 'Mod Transition Count'
        sheet[f'M{28 + i*13}'] = f'{start} - {end}'
        sheet[f'M{29 + i*13}'] = 'Count'
        sheet[f'N{28 + i*13}'] = 'To'

        sheet[f'N{29 + i*13}'] = sheet[f'M{30 + i*13}'] = '+1'
        sheet[f'O{29 + i*13}'] = sheet[f'M{31 + i*13}'] = '-1'
        sheet[f'P{29 + i*13}'] = sheet[f'M{32 + i*13}'] = '+2'
        sheet[f'Q{29 + i*13}'] = sheet[f'M{33 + i*13}'] = '-2'
        sheet[f'R{29 + i*13}'] = sheet[f'M{34 + i*13}'] = '+3'
        sheet[f'S{29 + i*13}'] = sheet[f'M{35 + i*13}'] = '-3'
        sheet[f'T{29 + i*13}'] = sheet[f'M{36 + i*13}'] = '+4'
        sheet[f'U{29 + i*13}'] = sheet[f'M{37 + i*13}'] = '-4'

        # Finding Transition count by comparing 2 columns, one of which contains one next values of the octant
        for j in range(8):
            sheet[f'N{30 + i*13 + j}'] = f'= COUNTIFS($K${start + 2}:$K${end + 2}, {cols[j]}, $AM${start + 2}:$AM${end + 2}, "+1")'
            sheet[f'O{30 + i*13 + j}'] = f'= COUNTIFS($K${start + 2}:$K${end + 2}, {cols[j]}, $AM${start + 2}:$AM${end + 2}, "-1")'
            sheet[f'P{30 + i*13 + j}'] = f'= COUNTIFS($K${start + 2}:$K${end + 2}, {cols[j]}, $AM${start + 2}:$AM${end + 2}, "+2")'
            sheet[f'Q{30 + i*13 + j}'] = f'= COUNTIFS($K${start + 2}:$K${end + 2}, {cols[j]}, $AM${start + 2}:$AM${end + 2}, "-2")'
            sheet[f'R{30 + i*13 + j}'] = f'= COUNTIFS($K${start + 2}:$K${end + 2}, {cols[j]}, $AM${start + 2}:$AM${end + 2}, "+3")'
            sheet[f'S{30 + i*13 + j}'] = f'= COUNTIFS($K${start + 2}:$K${end + 2}, {cols[j]}, $AM${start + 2}:$AM${end + 2}, "-3")'
            sheet[f'T{30 + i*13 + j}'] = f'= COUNTIFS($K${start + 2}:$K${end + 2}, {cols[j]}, $AM${start + 2}:$AM${end + 2}, "+4")'
            sheet[f'U{30 + i*13 + j}'] = f'= COUNTIFS($K${start + 2}:$K${end + 2}, {cols[j]}, $AM${start + 2}:$AM${end + 2}, "-4")'

    
    # Finding overall Transition count by comparing 2 columns same as above
    for j in range(8):
        sheet[f'N{16 + j}'] = f'= COUNTIFS($K$2:$K${row_count}, {cols[j]}, $AM$2:$AM${row_count}, "+1")'
        sheet[f'O{16 + j}'] = f'= COUNTIFS($K$2:$K${row_count}, {cols[j]}, $AM$2:$AM${row_count}, "-1")'
        sheet[f'P{16 + j}'] = f'= COUNTIFS($K$2:$K${row_count}, {cols[j]}, $AM$2:$AM${row_count}, "+2")'
        sheet[f'Q{16 + j}'] = f'= COUNTIFS($K$2:$K${row_count}, {cols[j]}, $AM$2:$AM${row_count}, "-2")'
        sheet[f'R{16 + j}'] = f'= COUNTIFS($K$2:$K${row_count}, {cols[j]}, $AM$2:$AM${row_count}, "+3")'
        sheet[f'S{16 + j}'] = f'= COUNTIFS($K$2:$K${row_count}, {cols[j]}, $AM$2:$AM${row_count}, "-3")'
        sheet[f'T{16 + j}'] = f'= COUNTIFS($K$2:$K${row_count}, {cols[j]}, $AM$2:$AM${row_count}, "+4")'
        sheet[f'U{16 + j}'] = f'= COUNTIFS($K$2:$K${row_count}, {cols[j]}, $AM$2:$AM${row_count}, "-4")'

    try:
        # Saving the Output file
        wb.save("output_octant_transition_identify.xlsx")
        # Above line can have error if the workbook has already been saved
        # and we are trying to save it again

    except Exception as e:  # Exception = openpyxl.utils.exceptions.WorkbookAlreadySaved
        print("There was some error due to " + str(e))

ver = python_version()

if ver == "3.8.10":
    print("Correct Version Installed")
else:
    print("Please install 3.8.10. Instruction are present in the GitHub Repo/Webmail. Url: https://pastebin.com/nvibxmjw")

try:
    mod = 5000
    # Error if mod is not an integer
except Exception as e:
    print("There was some error due to " + str(e))

octant_transition_count(mod)