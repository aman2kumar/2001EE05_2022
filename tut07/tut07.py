# Importing different modules
from platform import python_version

from msilib.schema import Font
from platform import python_version
# A Python library to read/write Excel file
import openpyxl as op
# For changing the style of a particular cell like changing colours, fonts or apply borders
from openpyxl.styles import PatternFill, Font, Border, Side

# Pandas is used for working on datasets
# It is simpler and easy to use
import pandas as pd

# import OS module
import os
# DateTime module to measure the execution time of the program
from datetime import datetime
start_time = datetime.now()

# Function to find the octant value and label it for a particular row
# taken row as an input
def label_octant (row):
    if row["U' = U - U_avg"] >= 0 and row["V' = V - V_avg"] >= 0 and row["W' = W - W_avg"] >= 0:
        return 1
    elif row["U' = U - U_avg"] < 0 and row["V' = V - V_avg"] >= 0 and row["W' = W - W_avg"] >= 0:
        return 2
    elif row["U' = U - U_avg"] < 0 and row["V' = V - V_avg"] < 0 and row["W' = W - W_avg"] >= 0:
        return 3
    elif row["U' = U - U_avg"] >= 0 and row["V' = V - V_avg"] < 0 and row["W' = W - W_avg"] >= 0:
        return 4
    elif row["U' = U - U_avg"] >= 0 and row["V' = V - V_avg"] >= 0 and row["W' = W - W_avg"] < 0:
        return -1
    elif row["U' = U - U_avg"] < 0 and row["V' = V - V_avg"] >= 0 and row["W' = W - W_avg"] < 0:
        return -2
    elif row["U' = U - U_avg"] < 0 and row["V' = V - V_avg"] < 0 and row["W' = W - W_avg"] < 0:
        return -3
    elif row["U' = U - U_avg"] >= 0 and row["V' = V - V_avg"] < 0 and row["W' = W - W_avg"] < 0:
        return -4

# Get the list of all files in input folder
dir_list = os.listdir("input")

def octant_analysis(mod = 5000): 
    for file in dir_list:

        try:
            # Loading input file into a dataframe (df)
            #file = "input/1.0.xlsx"
            df = pd.read_excel(f"input/{file}")
            # Above line can have error if input file has wrong format
            # other than .xlsx

        except Exception as e:  # Exception = openpyxl.utils.exceptions.InvalidFileException
            print("There was some error due to " + str(e))


        # Calculating averages of U, V and W
        U_avg = df["U"].mean()
        V_avg = df["V"].mean()
        W_avg = df["W"].mean()

        # Adding Columns of respective averages calculated above in first row
        df.loc[df.index[0], 'U Avg'] = U_avg
        df.loc[df.index[0], 'V Avg'] = V_avg
        df.loc[df.index[0], 'W Avg'] = W_avg

        # Done preprocessing by subtracting averages from original U, V and W
        # and forming new columns respectively
        df["U' = U - U_avg"] = df["U"] - U_avg
        df["V' = V - V_avg"] = df["V"] - V_avg
        df["W' = W - W_avg"] = df["W"] - W_avg

        # Constructing a new column of Octant and applying logic given below
        # Using lambda function on a particular row to find its octant via pre defined function (label_octant())
        df["Octant"] = df.apply(lambda row: label_octant(row), axis = 1)

        # Adding an empty column as shown in the given output file
        df[" "] = ""
        df[""] = ""


        # Calculating total rows
        total_rows = len(df.axes[0])

        octant_name_id_mapping = {"1":"Internal outward interaction", "-1":"External outward interaction", "2":"External Ejection", "-2":"Internal Ejection", 
        "3":"External inward interaction", "-3":"Internal inward interaction", "4":"Internal sweep", "-4":"External sweep"}

        
        # Adding Column of User Input at respective value of row and column
        df.at[2, ''] = 'Mod ' + str(mod)

        # Calculated total rows in dataframe to get the iteration of the for loop
        total_rows = len(df.axes[0])
        iteration = int(total_rows/mod)

        octant = [1, -1, 2, -2, 3, -3, 4, -4]
        overall_count = []

        # Added the total count of different octant values in their respective columns
        # ex- count_1 = len of dataframe df_range whose Octant value == 1
        # axes[0] => 0 corresponds to rows
        df.at[1, 'Octant ID'] = 'Overall Count'
        for i in range(8):
            count = len((df[df["Octant"] == octant[i]]).axes[0])
            df.at[1, str(octant[i])] = count
            overall_count.append([count, i + 1])

        for i in range(8):
            df.at[0, 'Rank ' + str(octant[i])] = ""

        overall_count.sort(reverse = True)
        for i in range(8):
            for j in range(8):
                if overall_count[j][1] == i + 1:
                    df.at[1, 'Rank ' + str(octant[i])] = j + 1
                    df.at[1, 'Rank 1 Octant ID'] = octant[overall_count[0][1] - 1]
                    df.at[1, 'Rank 1 Octant Name'] = octant_name_id_mapping.get(str(octant[overall_count[0][1] - 1]))
                    break

        # For loop to give the count of octants in the ranges of given mod value
        for i in range(iteration + 1):

            # Divided original dataframe in range based on no. of loops
            # ex- for loop1 range = 0 to 5000, loop2 range = 5001 to 10000 and so on..
            df_range = df[i*mod: (i + 1)*mod - 1]
            mod_count = []

            for j in range(8):
                count_mod = len((df_range[df_range["Octant"] == octant[j]]).axes[0])
                df.at[i + 2, str(octant[j])] = count_mod
                mod_count.append([count_mod, j + 1])

            mod_count.sort(reverse = True)
            for j in range(8):
                for k in range(8):
                    if mod_count[k][1] == j + 1:
                        df.at[i + 2, 'Rank ' + str(octant[j])] = k + 1
                        df.at[i + 2, 'Rank 1 Octant ID'] = octant[mod_count[0][1] - 1]
                        df.at[i + 2, 'Rank 1 Octant Name'] = octant_name_id_mapping.get(str(octant[mod_count[0][1] - 1]))   
                        break
        
            # Defined range based on loop value
            # For last loop range = i*mod to total_rows else shown below
            if i != int(iteration):
                df.at[i + 2, 'Octant ID'] = str(i*mod) + ' - ' + str((i+1)*mod - 1)
            else:
                df.at[i + 2, 'Octant ID'] = str(i*mod) + ' - ' + str(total_rows)

        df.at[iteration + 7, '1'] = 'Octant ID'
        df.at[iteration + 7, '-1'] = 'Octant Name'
        df.at[iteration + 7, '2'] = 'Count of Rank 1 Mod Values'
        for i in range(8):
            df.at[iteration + 8 + i, '1'] = octant[i]
            df.at[iteration + 8 + i, '-1'] = octant_name_id_mapping.get(str(octant[i]))
            if(i == 2):
                df.at[iteration + 8 + i, '2'] = len((df[df["Rank 1 Octant ID"] == octant[i]]).axes[0]) - 1
            else:
                df.at[iteration + 8 + i, '2'] = len((df[df["Rank 1 Octant ID"] == octant[i]]).axes[0])

        space_list = ["  ", "   ", "    ", "     ", "      ", "       ", "        ", "         ", "          ", "           ", "            ", "             "]
        for i in range(len(space_list)):
            df[space_list[i]] = ""
        
        # Defining Lists for different values of subsequence length for their respective octant values
        octant_values = [+1, -1, +2, -2, +3, -3, +4, -4]
        subsequence_length = [0, 0, 0, 0, 0, 0, 0, 0]
        max_subsequence_length = [-1, -1, -1, -1, -1, -1, -1, -1]
        temp_subsequence_length = [0, 0, 0, 0, 0, 0, 0, 0]
        max_subsequence_count = [0, 0, 0, 0, 0, 0, 0, 0]
        temp_start_time = [0, 0, 0, 0, 0, 0, 0, 0]
        temp_end_time = [0, 0, 0, 0, 0, 0, 0, 0]

        # Defining empty list of lists so that we can append it later
        # when we have multiple instances of longest subsequence length
        start_times = [[], [], [], [], [], [], [], []]
        end_times = [[], [], [], [], [], [], [], []]

        # Iteration for each row
        for i in range(total_rows - 1):

            # Iteration for each octant values
            for j in range(8):

                # If the value of Octant at index i == any octant values in the List
                if(df.loc[df.index[i], 'Octant'] == octant_values[j]):
                    # Then check for next octant value
                    if(df.loc[df.index[i + 1], 'Octant'] == octant_values[j] and i != total_rows - 2):
                        # if it is equal to previous octant value
                        # increment subsequence length by 1
                        subsequence_length[j] += 1
                    else:
                        # else set subsequence length to 0
                        subsequence_length[j] = 0

                    # define max subseq length as the maximum of the two
                    max_subsequence_length[j] = max(max_subsequence_length[j], subsequence_length[j])
                    break

        # Another loop to calculate max subsequence count and update start_times and end_times list
        for i in range(total_rows - 1):
            # Iteration for each octant values
            for j in range(8):
                # If the value of Octant at index i == any octant values in the List
                if(df.loc[df.index[i], 'Octant'] == octant_values[j]):
                    # Initialize defined variables
                    temp_start_time[j] = df.loc[df.index[i], 'T']
                    temp_subsequence_length[j] = 0
                    # while the next Octant value == to the starting one
                    while(df.loc[df.index[i], 'Octant'] == octant_values[j] and i != total_rows - 2):
                        # Increase the subsequence length
                        temp_subsequence_length[j] += 1
                        i += 1
                    temp_end_time[j] = df.loc[df.index[i - 1], 'T']
                    # Condition to check if current subsequence length == maximum length calculated earlier
                    if(temp_subsequence_length[j] == max_subsequence_length[j] + 1):
                        max_subsequence_count[j] += 1
                        start_times[j].append(temp_start_time[j])
                        end_times[j].append(temp_end_time[j])
                    break


        # Printing the calculated max suseq length and its count in the respective cells
        for i in range(8):
            df.loc[df.index[i], 'Octant No'] = str(octant_values[i])
            # Added 1 to the value because a single occurence also has the length 1
            df.loc[df.index[i], 'Longest Subsequence Length'] = max_subsequence_length[i] + 1
            df.loc[df.index[i], 'Count'] = max_subsequence_count[i]

        # Printing the time range according to the given format
        df["               "] = ""
        df.loc[df.index[0], 'Octant No '] = str(octant_values[0])
        df.loc[df.index[1], 'Octant No '] = 'Time'
        df.loc[df.index[0], 'Longest Subsequence Length '] = max_subsequence_length[0] + 1
        df.loc[df.index[1], 'Longest Subsequence Length '] = 'From'
        df.loc[df.index[2], 'Longest Subsequence Length '] = start_times[0][0]
        df.loc[df.index[0], 'Count '] = max_subsequence_count[0]
        df.loc[df.index[1], 'Count '] = 'To'
        df.loc[df.index[2], 'Count '] = end_times[0][0]

        # Below code is used to just format the excel file
        # doesn't contribute to the logic
        # by creating a list which stores max count till the that octant value
        indent = max_subsequence_count.copy()
        for i in range(1, 8):
            indent[i] = indent[i] + indent[i - 1]


        # Printing the time range according to the given format for different octant values
        for i in range(1, 8):
            df.loc[df.index[i*2 + indent[i - 1]], 'Octant No '] = str(octant_values[i])
            df.loc[df.index[i*2 + indent[i - 1] + 1], 'Octant No '] = 'Time'
            df.loc[df.index[i*2 + indent[i - 1]], 'Longest Subsequence Length '] = max_subsequence_length[i] + 1
            df.loc[df.index[i*2 + indent[i - 1] + 1], 'Longest Subsequence Length '] = 'From'
            for j in range(len(start_times[i])):
                df.loc[df.index[i*2 + indent[i - 1] + 2 + j], 'Longest Subsequence Length '] = start_times[i][j]
            df.loc[df.index[i*2 + indent[i - 1]], 'Count '] = max_subsequence_count[i]
            df.loc[df.index[i*2 + indent[i - 1] + 1], 'Count '] = 'To'
            for j in range(len(end_times[i])):
                df.loc[df.index[i*2 + indent[i - 1] + 2 + j], 'Count '] = end_times[i][j]

        try:
            # Forming excel file from the dataframe
            df.to_excel("demofile.xlsx", index = False)
            # Above line can have error if the workbook has already been saved

        except Exception as e:  # Exception
            print("There was some error due to " + str(e))

        try:
            # Loading the given input file in a Workbook
            wb = op.load_workbook('demofile.xlsx')
            # Above line can have error if input file has wrong format
            # other than .xlsx

        except Exception as e:  # Exception = openpyxl.utils.exceptions.InvalidFileException
            print("There was some error due to " + str(e))

        try:
            # Selecting the current active sheet
            sheet = wb.active
            # Above line can have error if input file is read only workbook
            # and we are trying to modify it

        except Exception as e:  # Exception = openpyxl.utils.exceptions.ReadOnlyWorkbookException
            print("There was some error due to " + str(e))

        # Defining Border type
        thin_border = Border(left = Side(style = 'thin'), right = Side(style = 'thin'), 
                                top = Side(style = 'thin'), bottom = Side(style = 'thin'))

        # Total rows
        row_count = sheet.max_row

        try:
            # Creating a temporary column containg one next values of octant
            sheet['BM1'] = "Temp_Column"
            for i in range(row_count):
                sheet[f'BM{i + 2}'] = f'= K{i + 3}'

        except:  # Exception = Unknown
            print("An Unknown error ocurred")


        # Creating List of cols value which will be used later
        cols = ['+1', '-1', '+2', '-2', '+3', '-3', '+4', '-4']

        # No of iterations for the below loop
        iter = row_count//mod
        sheet[f'AH4'] = 'From'
        sheet[f'AI1'] = 'Overall Transition Count'
        sheet[f'AI3'] = 'Octant #'
        sheet[f'AJ2'] = 'To'

        sheet[f'AJ3'] = sheet[f'AI4'] = '+1'
        sheet[f'AK3'] = sheet[f'AI5'] = '-1'
        sheet[f'AL3'] = sheet[f'AI6'] = '+2'
        sheet[f'AM3'] = sheet[f'AI7'] = '-2'
        sheet[f'AN3'] = sheet[f'AI8'] = '+3'
        sheet[f'AO3'] = sheet[f'AI9'] = '-3'
        sheet[f'AP3'] = sheet[f'AI10'] = '+4'
        sheet[f'AQ3'] = sheet[f'AI11'] = '-4'
        for i in range(iter + 1):
            # start = starting value of a given range
            # end = ending value of a given range
            start = i*mod
            if(i == iter):
                end = row_count
            else:
                end = (i + 1)*mod - 1

            # Assigning cell values for Transition count tables for each iteration
            sheet[f'AH{18 + i*13}'] = 'From'
            sheet[f'AI{15 + i*13}'] = 'Mod Transition Count'
            sheet[f'AI{16 + i*13}'] = f'{start} - {end}'
            sheet[f'AI{17 + i*13}'] = 'Octant #'
            sheet[f'AJ{16 + i*13}'] = 'To'

            sheet[f'AJ{17 + i*13}'] = sheet[f'AI{18 + i*13}'] = '+1'
            sheet[f'AK{17 + i*13}'] = sheet[f'AI{19 + i*13}'] = '-1'
            sheet[f'AL{17 + i*13}'] = sheet[f'AI{20 + i*13}'] = '+2'
            sheet[f'AM{17 + i*13}'] = sheet[f'AI{21 + i*13}'] = '-2'
            sheet[f'AN{17 + i*13}'] = sheet[f'AI{22 + i*13}'] = '+3'
            sheet[f'AO{17 + i*13}'] = sheet[f'AI{23 + i*13}'] = '-3'
            sheet[f'AP{17 + i*13}'] = sheet[f'AI{24 + i*13}'] = '+4'
            sheet[f'AQ{17 + i*13}'] = sheet[f'AI{25 + i*13}'] = '-4'
            
            # Finding Transition count by comparing 2 columns, one of which contains one next values of the octant
            for j in range(8):
                sheet[f'AJ{18 + i*13 + j}'] = f'= COUNTIFS($K${start + 2}:$K${end + 2}, {cols[j]}, $BM${start + 2}:$BM${end + 2}, "+1")'
                sheet[f'AK{18 + i*13 + j}'] = f'= COUNTIFS($K${start + 2}:$K${end + 2}, {cols[j]}, $BM${start + 2}:$BM${end + 2}, "-1")'
                sheet[f'AL{18 + i*13 + j}'] = f'= COUNTIFS($K${start + 2}:$K${end + 2}, {cols[j]}, $BM${start + 2}:$BM${end + 2}, "+2")'
                sheet[f'AM{18 + i*13 + j}'] = f'= COUNTIFS($K${start + 2}:$K${end + 2}, {cols[j]}, $BM${start + 2}:$BM${end + 2}, "-2")'
                sheet[f'AN{18 + i*13 + j}'] = f'= COUNTIFS($K${start + 2}:$K${end + 2}, {cols[j]}, $BM${start + 2}:$BM${end + 2}, "+3")'
                sheet[f'AO{18 + i*13 + j}'] = f'= COUNTIFS($K${start + 2}:$K${end + 2}, {cols[j]}, $BM${start + 2}:$BM${end + 2}, "-3")'
                sheet[f'AP{18 + i*13 + j}'] = f'= COUNTIFS($K${start + 2}:$K${end + 2}, {cols[j]}, $BM${start + 2}:$BM${end + 2}, "+4")'
                sheet[f'AQ{18 + i*13 + j}'] = f'= COUNTIFS($K${start + 2}:$K${end + 2}, {cols[j]}, $BM${start + 2}:$BM${end + 2}, "-4")'

        # Finding overall Transition count by comparing 2 columns same as above
        for j in range(8):
            sheet[f'AJ{4 + j}'] = f'= COUNTIFS($K$2:$K${row_count}, {cols[j]}, $BM$2:$BM${row_count}, "+1")'
            sheet[f'AK{4 + j}'] = f'= COUNTIFS($K$2:$K${row_count}, {cols[j]}, $BM$2:$BM${row_count}, "-1")'
            sheet[f'AL{4 + j}'] = f'= COUNTIFS($K$2:$K${row_count}, {cols[j]}, $BM$2:$BM${row_count}, "+2")'
            sheet[f'AM{4 + j}'] = f'= COUNTIFS($K$2:$K${row_count}, {cols[j]}, $BM$2:$BM${row_count}, "-2")'
            sheet[f'AN{4 + j}'] = f'= COUNTIFS($K$2:$K${row_count}, {cols[j]}, $BM$2:$BM${row_count}, "+3")'
            sheet[f'AO{4 + j}'] = f'= COUNTIFS($K$2:$K${row_count}, {cols[j]}, $BM$2:$BM${row_count}, "-3")'
            sheet[f'AP{4 + j}'] = f'= COUNTIFS($K$2:$K${row_count}, {cols[j]}, $BM$2:$BM${row_count}, "+4")'
            sheet[f'AQ{4 + j}'] = f'= COUNTIFS($K$2:$K${row_count}, {cols[j]}, $BM$2:$BM${row_count}, "-4")'

        highlight_columns = ["W", "X", "Y", "Z", "AA", "AB", "AC", "AD"]
        for i in range(iter + 2):
            for col in highlight_columns:
                if sheet[f"{col}{i + 3}"].value == 1:
                    # Fills the selected colour in the chosen cell
                    sheet[f"{col}{i + 3}"].fill = PatternFill(start_color = 'FFFF00', end_color = 'FFFF00', fill_type = 'solid')


        # Below codes are to add borders in respective group of cells
        overall_count_columns = ["N", "O", "P", "Q", "R", "S", "T", "U", "V", "W", "X", "Y", "Z", "AA", "AB", "AC", "AD", "AE", "AF"]
        for col in overall_count_columns:
            range_cell = sheet[f'{col}{3}':f'{col}{iter + 4}']
            for cell in range_cell:
                for x in cell:
                    x.border = thin_border
        
        transition_count_column = ["AI", "AJ", "AK", "AL", "AM", "AN", "AO", "AP", "AQ"]
        for col in transition_count_column:
            range_cell = sheet[f'{col}{3}':f'{col}{11}']
            for cell in range_cell:
                for x in cell:
                    x.border = thin_border
        for i in range(iter + 1):
            for col in transition_count_column:
                range_cell = sheet[f'{col}{17 + i*13}':f'{col}{25 + i*13}']
                for cell in range_cell:
                    for x in cell:
                        x.border = thin_border


        longest_sub_column = ["AS", "AT", "AU"]
        for col in longest_sub_column:
            range_cell = sheet[f'{col}{1}':f'{col}{9}']
            for cell in range_cell:
                for x in cell:
                    x.border = thin_border

        longest_sub_wrange_column = ["AW", "AX", "AY"]
        for col in longest_sub_wrange_column:
            range_cell = sheet[f'{col}{1}':f'{col}{25}']
            for cell in range_cell:
                for x in cell:
                    x.border = thin_border
        
        below_column = ["O", "P", "Q"]
        for col in below_column:
            range_cell = sheet[f'{col}{12}':f'{col}{20}']
            for cell in range_cell:
                for x in cell:
                    x.border = thin_border

        file = file[:-5]
        try:
            # Saving the Output file in output folder
            wb.save(f"output/{file}_vel_octant_analysis_mod_{mod}.xlsx")
            # Above line can have error if the workbook has already been saved
            # and we are trying to save it again

        except Exception as e:  # Exception = openpyxl.utils.exceptions.WorkbookAlreadySaved
            print("There was some error due to " + str(e))

        os.remove("demofile.xlsx")



ver = python_version()

if ver == "3.8.10":
	print("Correct Version Installed")
else:
	print("Please install 3.8.10. Instruction are present in the GitHub Repo/Webmail. Url: https://pastebin.com/nvibxmjw")


mod = 5000
octant_analysis(mod)

#This shall be the last lines of the code.
end_time = datetime.now()
print('Duration of Program Execution: {}'.format(end_time - start_time))
